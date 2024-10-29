import pandas as pd
import PySimpleGUI as sg
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import json
import traceback
import re
import numpy as np
import warnings
import time

# Define DEFAULT_REGEX_EXPRESSIONS globally
DEFAULT_REGEX_EXPRESSIONS = [
    ('(?i)pattern', 'Case-insensitive pattern'),
    ('pattern', 'Case-sensitive pattern'),
    (r'\d+', 'One or more digits'),
    (r'\w+', 'One or more word characters'),
    (r'\s+', 'One or more whitespace characters'),
    # Add more default patterns as needed
]

# Add this near the top of your file, after the imports
color_categories = [
    ("#FFFF00", "Yellow", "keyword1, keyword2"),
    ("#E6E6FA", "Lavender", "keyword3, keyword4"),
    ("#90EE90", "Light Green", "keyword5, keyword6"),
    ("#ADD8E6", "Light Blue", "keyword7, keyword8"),
    ("#FFB6C1", "Light Pink", "keyword9, keyword10"),
    ("#FFA500", "Orange", "keyword11, keyword12"),
    ("#00CED1", "Dark Turquoise", "keyword13, keyword14"),
    ("#FF69B4", "Hot Pink", "keyword15, keyword16")
]

print("Script started")

def add_new_records(df, file_path):
    print("add_new_records function called")
    
    background_color = '#0C2340'
    text_color = '#FFFFFF'
    button_color = ('#FFFFFF', '#C4122F')
    input_background_color = '#F0F0F0'
    
    field_lengths = {
        'NUMBER': 10, 'DWG': 15, 'ORIGIN': 30, 'DEST': 30,
        'Alt DWG': 15, 'Wire Type': 20, 'Length': 10,
        'Note': 40, 'Project ID': 20
    }
    
    layout = [
        [sg.Text("Add New Records", font=('Helvetica', 16), text_color=text_color, background_color=background_color)],
    ]
    
    for col in ['NUMBER', 'DWG', 'ORIGIN', 'DEST', 'Alt DWG', 'Wire Type', 'Length', 'Note', 'Project ID']:
        if col in df.columns:
            layout.append([
                sg.Text(col, size=(10, 1), text_color=text_color, background_color=background_color),
                sg.Input(key=f"-NEW-{col}-", size=(field_lengths[col], 1), background_color=input_background_color)
            ])
    
    layout.extend([
        [sg.Button("Add Record", button_color=button_color, bind_return_key=True), 
         sg.Button("Save All", button_color=button_color),
         sg.Button("Cancel", button_color=button_color)]
    ])
    
    window = sg.Window("Add New Records", layout, background_color=background_color, finalize=True, return_keyboard_events=True)
    
    new_records = []
    
    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, "Cancel", "\r"):  # \r is the Return key
            break
        
        if event in ("Add Record", "\r"):  # Handle both button click and Return key
            new_record = {col: values.get(f"-NEW-{col}-", "") for col in df.columns}
            try:
                new_record['NUMBER'] = int(new_record['NUMBER'])
                new_records.append(new_record)
                window["-RECORDS-"].print(f"Added: {new_record}")
                for key in values:
                    if key.startswith("-NEW-"):
                        window[key].update("")
            except ValueError:
                sg.popup_error("NUMBER must be an integer.", background_color=background_color, text_color=text_color)
        
        if event == "Save All" and new_records:
            df_new = pd.DataFrame(new_records)
            df = pd.concat([df, df_new], ignore_index=True)
            try:
                with pd.ExcelFile(file_path) as xls:
                    other_sheets = {sheet_name: pd.read_excel(xls, sheet_name)
                                    for sheet_name in xls.sheet_names if sheet_name != "CableList"}
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="CableList", index=False)
                    for sheet_name, sheet_df in other_sheets.items():
                        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                sg.popup(f"{len(new_records)} new records added successfully!", background_color=background_color, text_color=text_color)
                break
            except Exception as e:
                sg.popup_error(f"Error adding new records: {str(e)}", background_color=background_color, text_color=text_color)
    
    window.close()
    return df

def load_data(file_path):
    print(f"Loading data from {file_path}")
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        if not file_path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError("Invalid file type. Please select an Excel file (.xlsx or .xlsm)")

        with pd.ExcelFile(file_path) as xls:
            print(f"Excel file opened. Sheets: {xls.sheet_names}")
            required_sheets = ["CableList", "LengthMatrix"]
            missing_sheets = [sheet for sheet in required_sheets if sheet not in xls.sheet_names]
            if missing_sheets:
                raise ValueError(f"Missing required sheets: {', '.join(missing_sheets)}")

            cable_list = pd.read_excel(xls, sheet_name="CableList")
            length_matrix = pd.read_excel(xls, sheet_name="LengthMatrix", index_col=0)

        print(f"CableList shape: {cable_list.shape}")
        print(f"LengthMatrix shape: {length_matrix.shape}")

        required_columns = ['NUMBER', 'DWG', 'ORIGIN', 'DEST', 'Wire Type', 'Length', 'Note', 'Project ID']
        missing_columns = [col for col in required_columns if col not in cable_list.columns]
        
        if missing_columns:
            raise ValueError(f"Missing required columns in CableList: {', '.join(missing_columns)}")

        cable_list['NUMBER'] = pd.to_numeric(cable_list['NUMBER'], errors='coerce').astype('Int64')

        print("Data loaded successfully.")
        return cable_list, length_matrix
    except Exception as e:
        print(f"Error loading data: {str(e)}")
        traceback.print_exc()
        return None, None

def apply_filter(df, values):
    filtered_df = df.copy()
    
    # NUMBER filter
    number_single = values.get("-NUMBER_SINGLE-")
    number_range = values.get("-NUMBER_RANGE-")
    if number_single:
        try:
            filtered_df = filtered_df[filtered_df['NUMBER'] == int(number_single)]
        except ValueError:
            sg.popup_error("Invalid NUMBER input")
    elif number_range:
        try:
            start, end = map(int, number_range.split('-'))
            filtered_df = filtered_df[(filtered_df['NUMBER'] >= start) & (filtered_df['NUMBER'] <= end)]
        except ValueError:
            sg.popup_error("Invalid NUMBER range")

    # Other column filters
    for col in df.columns:
        if col != 'NUMBER':
            filter_value = values.get(f"-{col}-")
            exact_match = values.get(f"-{col}-EXACT-")
            if filter_value:
                if exact_match:
                    filtered_df = filtered_df[filtered_df[col].astype(str) == filter_value]
                else:
                    filtered_df = filtered_df[filtered_df[col].astype(str).str.contains(filter_value, case=False, na=False)]
    
    return filtered_df

def create_rack_mapping(length_matrix):
    rack_mapping = {}
    for rack in length_matrix.index:
        rack_mapping[rack] = rack
        if rack.startswith('T'):
            rack_mapping[rack[1:]] = rack  # Map "G01" to "TG01"
    return rack_mapping

def get_rack(name, mapping):
    return mapping.get(name[:4])  # Assume rack names are in the first 4 characters

def update_lengths_from_matrix(df, length_matrix, start_number, end_number, rack_mapping):
    print(f"Entering update_lengths_from_matrix. DataFrame shape: {df.shape}")
    mask = (df['NUMBER'] >= start_number) & (df['NUMBER'] <= end_number) & df['Length'].isna()
    df_to_update = df.loc[mask]
    print(f"Rows to update: {len(df_to_update)}")
    
    def get_length(row):
        origin_rack = get_rack(str(row['ORIGIN']), rack_mapping)
        dest_rack = get_rack(str(row['DEST']), rack_mapping)
        if origin_rack and dest_rack and origin_rack in length_matrix.index and dest_rack in length_matrix.columns:
            return length_matrix.loc[origin_rack, dest_rack]
        return np.nan
    
    df.loc[mask, 'Length'] = df_to_update.apply(get_length, axis=1)
    print(f"Updated DataFrame shape: {df.shape}")
    return df

def load_favorite_colors():
    try:
        with open('favorite_colors.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return ["#FFFF00", "#E6E6FA", "#90EE90", "#ADD8E6", "#FFB6C1", "#FFA500", "#00CED1", "#FF69B4"]

def save_favorite_colors(colors):
    with open('favorite_colors.json', 'w') as f:
        json.dump(colors, f)

def load_saved_regex():
    try:
        with open('saved_regex.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_regex(regex_dict):
    with open('saved_regex.json', 'w') as f:
        json.dump(regex_dict, f)

def update_table(window, df, primary_color, secondary_color, text_color):
    visible_columns = [col for col in df.columns if col not in ['NUMC'] + [f'F{i}' for i in range(11, 22)]]
    data = df[visible_columns].values.tolist()
    window["-TABLE-"].update(values=data)
    
    # Prepare display data
    display_data = []
    for _, row in df.iterrows():
        row_data = []
        for col in df.columns:
            value = row[col]
            if pd.isna(value):
                row_data.append('')
            elif df[col].dtype == 'Int64':
                row_data.append(str(value) if pd.notna(value) else '')
            else:
                row_data.append(str(value))
        display_data.append(row_data)
    
    print(f"Prepared display data. Rows: {len(display_data)}, Columns: {len(display_data[0]) if display_data else 0}")
    
    table = window["-TABLE-"]
    
    # Get current selections
    selected_rows = table.SelectedRows if hasattr(table, 'SelectedRows') else []
    print(f"Current selected rows: {selected_rows}")
    
    # Create a list of row colors
    row_colors = [primary_color, secondary_color] * (len(display_data) // 2 + 1)
    row_color_list = [(i, text_color, color) for i, color in enumerate(row_colors[:len(display_data)])]
    
    print("Updating table...")
    # Update table values and colors
    table.update(values=display_data, row_colors=row_color_list)
    
    # Reapply row selection if any
    if selected_rows:
        print(f"Reapplying row selection: {selected_rows}")
        table.update(select_rows=selected_rows)
    
    window.refresh()
    print("Table updated and window refreshed")
    
def create_layout(df, length_matrix_headers, add_new_records_func, background_color, text_color, button_color, input_background_color, color_categories):
    visible_columns = [col for col in df.columns if col not in ['NUMC'] + [f'F{i}' for i in range(11, 22)]]
    
    filter_layout = [
        [sg.Text("Filters", font=('Helvetica', 16), text_color=text_color, background_color=background_color)],
    ]

    # Add NUMBER filter
    filter_layout.extend([
        [sg.Text("NUMBER:", size=(10, 1), justification='right', background_color=background_color, text_color=text_color),
         sg.Input(key="-NUMBER_SINGLE-", size=(10,1), background_color=input_background_color),
         sg.Text("to", background_color=background_color, text_color=text_color),
         sg.Input(key="-NUMBER_RANGE-", size=(10,1), background_color=input_background_color)]
    ])

    # Add other filters
    for col in visible_columns:
        if col != 'NUMBER':
            filter_layout.extend([
                [sg.Text(f"{col}:", size=(10, 1), justification='right', background_color=background_color, text_color=text_color),
                 sg.Input(key=f"-{col}-", size=(20, 1), background_color=input_background_color),
                 sg.Checkbox('Exact', key=f"-{col}-EXACT-", background_color=background_color, text_color=text_color)]
            ])
    
    filter_layout.append([sg.Button("Apply Filter", button_color=button_color, bind_return_key=True), sg.Button("Clear Filter", button_color=button_color)])
    
    sort_group_layout = [
        [sg.Text("Sort By:", size=(10, 1), justification='right', background_color=background_color, text_color=text_color), 
         sg.Combo(visible_columns, key="-SORT-", background_color=input_background_color), 
         sg.Radio("Ascending", "SORT", default=True, key="-ASCENDING-", background_color=background_color), 
         sg.Radio("Descending", "SORT", key="-DESCENDING-", background_color=background_color),
         sg.Button("Sort", button_color=button_color)],
        [sg.Text("Group By:", size=(10, 1), justification='right', background_color=background_color, text_color=text_color), 
         sg.Radio("Origin", "GROUP", key="-GROUP_ORIGIN-", background_color=background_color), 
         sg.Radio("Destination", "GROUP", key="-GROUP_DEST-", background_color=background_color),
         sg.Combo(length_matrix_headers, key="-GROUP_VALUE-", background_color=input_background_color),
         sg.Button("Apply Grouping", button_color=button_color), sg.Button("Revert Grouping", button_color=button_color)]
    ]
    
    color_layout = [[sg.Text("Color Categories:", background_color=background_color, text_color=text_color)]]
    for i, (color_code, color_name, keywords) in enumerate(color_categories):
        color_layout.append([
            sg.Text(f"{color_name}:", size=(12,1), justification='right', background_color=background_color, text_color=text_color),
            sg.Input(default_text=color_code, size=(8,1), key=f"-COLOR_CODE_{i}-", background_color=input_background_color),
            sg.ColorChooserButton("Pick", target=f"-COLOR_CODE_{i}-", key=f"-COLOR_PICKER_{i}-", button_color=button_color),
            sg.Input(default_text=keywords, key=f"-COLOR_KEYWORDS_{i}-", size=(30,1), background_color=input_background_color)
        ])
    color_layout.append([sg.Button("Add Category", button_color=button_color)])
    
    col_widths = {
        'NUMBER': 5,
        'DWG': 8,
        'ORIGIN': 30,
        'DEST': 30,
        'Wire Type': 15,
        'Length': 10,
        'Note': 30,
        'Project ID': 20
    }
    
    button_column = [
        [sg.Button("LengthMatrix Lookup", button_color=('#FFFFFF', '#C4122F'))],
        [sg.Button("Save Formatted Excel", button_color=('#FFFFFF', '#C4122F'))],
        [sg.Button("Save Changes to Source", button_color=('#FFFFFF', '#C4122F'))],
        [sg.Button("Add New Record", key="-ADD_NEW_RECORD-", button_color=('#FFFFFF', '#C4122F'))],
        [sg.Button("Import CSV", key="-IMPORT_CSV-", button_color=('#FFFFFF', '#C4122F'))],
        [sg.Button("Reload Data", key="-RELOAD_DATA-", button_color=('#FFFFFF', '#C4122F'))],  # New button
        [sg.Button("Exit", button_color=('#FFFFFF', '#C4122F'))]
    ]
    print(f"Button column: {button_column}")  # Debug print
    
    edit_layout = [
        [sg.Text("Selected Cell:"),
         sg.Text("", size=(15, 1), key="-SELECTED-CELL-"),
         sg.Text("Edit Cell:"),
         sg.Input(key="-CELL-CONTENT-", size=(30, 1)),
         sg.Button("Update Cell")]
    ]
    
    # Prepare initial table data
    initial_data = []
    for _, row in df[visible_columns].iterrows():
        row_data = []
        for col in visible_columns:
            value = row[col]
            if pd.isna(value):
                row_data.append('')
            elif df[col].dtype == 'Int64':
                row_data.append(str(value) if pd.notna(value) else '')
            else:
                row_data.append(str(value))
        initial_data.append(row_data)

    table_layout = [
        [sg.Table(values=initial_data,
                  headings=visible_columns,
                  display_row_numbers=True,
                  auto_size_columns=False,
                  def_col_width=12,
                  col_widths=[col_widths.get(col, 12) for col in visible_columns],
                  num_rows=25,
                  key="-TABLE-",
                  enable_events=True,
                  expand_x=True,
                  expand_y=True,
                  text_color='#ECF0F1',  # Very light gray, almost white
                  font=('Any', 10, 'bold'))]
    ]
    
    settings_icon = '⚙️'  # Unicode gear emoji
    settings_button = sg.Button(settings_icon, key='-SETTINGS-', size=(2,1))

    layout = [
        [settings_button, sg.Push()],  # Add this at the top of your layout
        [sg.Column(filter_layout, vertical_alignment='top', background_color=background_color), 
         sg.VSeparator(), 
         sg.Column(color_layout, vertical_alignment='top', background_color=background_color),
         sg.VSeparator(),
         sg.Column(button_column, vertical_alignment='top', background_color=background_color)],
        [sg.HorizontalSeparator()],
        [sg.Column(sort_group_layout, background_color=background_color)],
        [sg.HorizontalSeparator()],
        table_layout
    ]
    
    print("Layout created")
    return layout
    
def save_formatted_excel(df, color_categories, output_file):
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Convert NUMBER to string to remove decimal places
            df['NUMBER'] = df['NUMBER'].astype('Int64').astype(str)
            df.fillna('').to_excel(writer, sheet_name='Formatted_CableList', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Formatted_CableList']
            
            color_fills = {color_code: PatternFill(start_color=color_code.lstrip('#'), end_color=color_code.lstrip('#'), fill_type="solid") 
                           for color_code, _, _ in color_categories}
            
            for row in range(2, len(df) + 2):  # Start from 2 to skip header
                cell_color = worksheet.cell(row=row, column=df.columns.get_loc("Color") + 1).value
                fill = color_fills.get(cell_color)
                
                if fill:
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = fill
            
            worksheet.delete_cols(df.columns.get_loc("Color") + 1)  # Remove the Color column
        return True
    except PermissionError:
        sg.popup_error("Permission denied. The file may be open in another program.")
        return False
    except Exception as e:
        sg.popup_error(f"An error occurred while saving the file: {str(e)}")
        return False
        
   

   
        
def save_changes_to_excel(df, file_path):
    try:
        # Read the existing Excel file
        with pd.ExcelFile(file_path) as xls:
            other_sheets = {sheet_name: pd.read_excel(xls, sheet_name)
                            for sheet_name in xls.sheet_names if sheet_name != "CableList"}
        
        # Write the updated dataframe and other sheets back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="CableList", index=False)
            for sheet_name, sheet_df in other_sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return True
    except Exception as e:
        print(f"Error saving changes: {str(e)}")
        return False




def apply_sort(df, sort_column, ascending):
    return df.sort_values(by=sort_column, ascending=ascending)




def apply_grouping(df, group_by, group_value):
    if group_by not in ["ORIGIN", "DEST"]:
        return df
    
    column = group_by
    
    if not group_value:
        unique_values = df[column].unique()
        return pd.DataFrame({column: unique_values})
    
    mask = df[column].str.lower().str.contains(group_value.lower(), na=False)
    return df[mask]
    
    
    
    
def color_code_rows(df, color_categories):
    def get_color(row):
        origin_dest = f"{row['ORIGIN']} {row['DEST']}"
        for color_code, _, keywords in color_categories:
            if any(keyword.strip().lower() in origin_dest.lower() for keyword in keywords.split(',')):
                return color_code
        return "FFFFFF"  # White
    
    df["Color"] = df.apply(get_color, axis=1)
    return df





def generate_default_filename(df):
    current_date = datetime.now().strftime("%Y%m%d")
    dwg = df['DWG'].iloc[0] if 'DWG' in df.columns else 'UnknownDWG'
    project_id = df['Project ID'].iloc[0] if 'Project ID' in df.columns else 'UnknownProject'
    return f"{dwg}_{project_id}_{current_date}.xlsx"





def save_formatted_excel(df, color_categories, output_file):
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Convert NUMBER to string to remove decimal places
            df['NUMBER'] = df['NUMBER'].astype('Int64').astype(str)
            df.fillna('').to_excel(writer, sheet_name='Formatted_CableList', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Formatted_CableList']
            
            color_fills = {color_code: PatternFill(start_color=color_code.lstrip('#'), end_color=color_code.lstrip('#'), fill_type="solid") 
                           for color_code, _, _ in color_categories}
            
            for row in range(2, len(df) + 2):  # Start from 2 to skip header
                cell_color = worksheet.cell(row=row, column=df.columns.get_loc("Color") + 1).value
                fill = color_fills.get(cell_color)
                
                if fill:
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = fill
            
            worksheet.delete_cols(df.columns.get_loc("Color") + 1)  # Remove the Color column
        return True
    except PermissionError:
        sg.popup_error("Permission denied. The file may be open in another program.")
        return False
    except Exception as e:
        sg.popup_error(f"An error occurred while saving the file: {str(e)}")
        return False




def manage_regex(column, current_regex, saved_regex):
    layout = [
        [sg.Text(f"Manage Regex for {column}")],
        [sg.Text("Current Regex:"), sg.Input(current_regex, key="-CURRENT_REGEX-", size=(30, 1))],
        [sg.Text("Saved Regex:"), sg.Listbox(list(saved_regex.get(column, {}).keys()), size=(30, 6), key="-SAVED_REGEX_LIST-")],
        [sg.Button("Save Current", bind_return_key=True), sg.Button("Apply Selected"), sg.Button("Delete Selected")],
        [sg.Button("Close")]
    ]
    
    window = sg.Window(f"Manage Regex for {column}", layout, finalize=True, return_keyboard_events=True)
    
    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Close", "\r"):  # \r is the Return key
            break
        elif event in ("Save Current", "\r"):  # Handle both button click and Return key
            regex = values["-CURRENT_REGEX-"]
            if regex:
                name = sg.popup_get_text(f"Enter a name for this {column} regex:")
                if name:
                    if column not in saved_regex:
                        saved_regex[column] = {}
                    saved_regex[column][name] = regex
                    save_regex(saved_regex)
                    window["-SAVED_REGEX_LIST-"].update(values=list(saved_regex[column].keys()))
                    print(f"Saved regex '{name}' for column '{column}': {regex}")
        elif event == "Apply Selected":
            selected = values["-SAVED_REGEX_LIST-"]
            if selected:
                selected_regex = selected[0]
                regex = saved_regex[column][selected_regex]
                window["-CURRENT_REGEX-"].update(regex)
                print(f"Applied regex '{selected_regex}' to column '{column}': {regex}")
        elif event == "Delete Selected":
            selected = values["-SAVED_REGEX_LIST-"]
            if selected:
                selected_regex = selected[0]
                del saved_regex[column][selected_regex]
                save_regex(saved_regex)
                window["-SAVED_REGEX_LIST-"].update(values=list(saved_regex[column].keys()))
                print(f"Deleted regex '{selected_regex}' from column '{column}'")
    
    window.close()
    return values["-CURRENT_REGEX-"]





def handle_length_matrix_lookup(df, length_matrix, window, rack_mapping):
    layout = [
        [sg.Text("Select range of cable numbers to update:")],
        [sg.Text("Start Number:", size=(12, 1), justification='right'), sg.Input(key="-START_NUMBER-", size=(6, 1))],
        [sg.Text("End Number:", size=(12, 1), justification='right'), sg.Input(key="-END_NUMBER-", size=(6, 1))],
        [sg.Button("Preview", bind_return_key=True), sg.Button("Cancel")]
    ]
    
    range_window = sg.Window("LengthMatrix Lookup", layout, finalize=True, return_keyboard_events=True)
    
    while True:
        event, values = range_window.read()
        print(f"LengthMatrix Lookup event: {event}")
        
        if event in (sg.WIN_CLOSED, "Cancel", "\r"):
            print("Exiting LengthMatrix Lookup without changes")
            break
        
        if event in ("Preview", "\r"):
            try:
                start_number = int(values["-START_NUMBER-"])
                end_number = int(values["-END_NUMBER-"])
                print(f"Start number: {start_number}, End number: {end_number}")
                
                if start_number > end_number:
                    raise ValueError("Invalid number range")
                
                preview_df = update_lengths_from_matrix(df.copy(), length_matrix, start_number, end_number, rack_mapping)
                print(f"Preview dataframe created. Shape: {preview_df.shape}")
                
                preview_mask = (preview_df['NUMBER'] >= start_number) & (preview_df['NUMBER'] <= end_number)
                preview_data = preview_df[preview_mask]
                print(f"Preview data created. Shape: {preview_data.shape}")

                preview_data_str = preview_data.astype(str).replace('nan', '')
                
                preview_layout = [
                    [sg.Table(values=preview_data_str.values.tolist(),
                              headings=preview_data.columns.tolist(),
                              display_row_numbers=False,
                              auto_size_columns=False,
                              num_rows=min(10, len(preview_data)),
                              key="-PREVIEW_TABLE-")],
                    [sg.Button("Confirm"), sg.Button("Cancel")]
                ]
                preview_window = sg.Window("Preview Changes", preview_layout)
                
                preview_event, _ = preview_window.read()
                print(f"Preview window event: {preview_event}")
                if preview_event == "Confirm":
                    df = update_lengths_from_matrix(df, length_matrix, start_number, end_number, rack_mapping)
                    print("LengthMatrix lookup completed successfully")
                    sg.popup("LengthMatrix lookup completed successfully!")
                    preview_window.close()
                    range_window.close()
                    return df
                
                preview_window.close()
            
            except ValueError as e:
                print(f"ValueError in LengthMatrix Lookup: {str(e)}")
                sg.popup_error(f"Error: {str(e)}")
            except Exception as e:
                print(f"Unexpected error in LengthMatrix Lookup: {str(e)}")
                sg.popup_error(f"An unexpected error occurred: {str(e)}")
    
    range_window.close()
    print("Exiting handle_length_matrix_lookup function")
    return df
    
    
    
    
    
  

















  

def show_loading_animation():
    layout = [[sg.Text('Loading...', font=('Helvetica', 16))],
              [sg.ProgressBar(100, orientation='h', size=(20, 20), key='progressbar')]]
    window = sg.Window('Loading', layout, finalize=True, keep_on_top=True, no_titlebar=True)
    progress_bar = window['progressbar']
    for i in range(100):
        event, values = window.read(timeout=10)
        if event == sg.WINDOW_CLOSED:
            break
        progress_bar.UpdateBar(i + 1)
    window.close()

def open_settings_drawer(current_settings):
    layout = [
        [sg.Text("Color Settings", font=('Helvetica', 16))],
        [sg.Text("Background Color:"), sg.Input(current_settings['background_color'], key='-BG_COLOR-'), sg.ColorChooserButton("Pick", target='-BG_COLOR-')],
        [sg.Text("Text Color:"), sg.Input(current_settings['text_color'], key='-TEXT_COLOR-'), sg.ColorChooserButton("Pick", target='-TEXT_COLOR-')],
        [sg.Text("Button Color:"), sg.Input(current_settings['button_color'][1], key='-BUTTON_COLOR-'), sg.ColorChooserButton("Pick", target='-BUTTON_COLOR-')],
        [sg.Text("Input Background:"), sg.Input(current_settings['input_background_color'], key='-INPUT_BG_COLOR-'), sg.ColorChooserButton("Pick", target='-INPUT_BG_COLOR-')],
        [sg.Button("Save"), sg.Button("Cancel")]
    ]

    window = sg.Window("Settings", layout, finalize=True, keep_on_top=True)

    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, "Cancel"):
            break
        if event == "Save":
            new_settings = {
                'background_color': values['-BG_COLOR-'],
                'text_color': values['-TEXT_COLOR-'],
                'button_color': (current_settings['button_color'][0], values['-BUTTON_COLOR-']),
                'input_background_color': values['-INPUT_BG_COLOR-']
            }
            window.close()
            return new_settings

    window.close()
    return None

def load_last_file_path():
    try:
        with open('last_file_path.json', 'r') as f:
            data = json.load(f)
            return data.get('last_path', '')
    except FileNotFoundError:
        return ''

def save_last_file_path(file_path):
    with open('last_file_path.json', 'w') as f:
        json.dump({'last_path': file_path}, f)

def load_settings():
    try:
        with open('settings.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        # Default settings
        return {
            'background_color': '#0C2340',
            'text_color': '#FFFFFF',
            'button_color': ('#FFFFFF', '#C4122F'),
            'input_background_color': '#F0F0F0'
        }

def save_settings(settings):
    with open('settings.json', 'w') as f:
        json.dump(settings, f)

def main():
    print("Entering main function")
    try:
        # Load last file path
        last_path = load_last_file_path()
        print(f"Last file path: {last_path}")

        if last_path and os.path.exists(last_path):
            file_path = last_path
            print(f"Existing file path found: {file_path}")
            layout = [
                [sg.Text(f"Load last used file?\n{file_path}")],
                [sg.Button("Yes", bind_return_key=True), sg.Button("No")]
            ]
            window = sg.Window("Load File", layout, finalize=True, return_keyboard_events=True)
            print("File load window created")
            event, _ = window.read()
            print(f"File load window event: {event}")
            window.close()
            
            if event in (sg.WIN_CLOSED, "No"):
                print("User chose not to load last file")
                file_path = sg.popup_get_file("Select the Excel file", file_types=(("Excel Files", "*.xlsm;*.xlsx"),), initial_folder=os.path.dirname(last_path))
        else:
            print("No existing file path or file not found")
            file_path = sg.popup_get_file("Select the Excel file", file_types=(("Excel Files", "*.xlsm;*.xlsx"),), initial_folder=os.path.dirname(last_path) if last_path else None)
        
        print(f"Selected file path: {file_path}")

        if not file_path:
            print("No file selected, exiting")
            return

        # Save the selected file path
        save_last_file_path(file_path)

        # Load data
        print("Loading data...")
        result = load_data(file_path)
        print(f"load_data returned: {type(result)}")
        
        if isinstance(result, tuple) and len(result) == 2:
            df, length_matrix = result
            print(f"df type: {type(df)}, length_matrix type: {type(length_matrix)}")
        else:
            print(f"Unexpected result from load_data: {result}")
            raise ValueError("Failed to load data. Please check the file and try again.")

        if df is None or length_matrix is None:
            raise ValueError("Failed to load data. Please check the file and try again.")

        print("Data loaded successfully")
        print(f"df shape: {df.shape}, length_matrix shape: {length_matrix.shape}")

        # Create rack_mapping
        rack_mapping = create_rack_mapping(length_matrix)

        # Load or initialize settings
        settings = load_settings()
        background_color = settings['background_color']
        text_color = settings['text_color']
        button_color = settings['button_color']
        input_background_color = settings['input_background_color']

        # Create layout and window
        layout = create_layout(df, length_matrix.columns.tolist(), add_new_records, 
                               background_color, text_color, button_color, input_background_color,
                               color_categories)
        window = sg.Window("Cable Database Interface", layout, resizable=True, finalize=True)
        window.maximize()

        print("Main window created")

        # Main event loop
        while True:
            event, values = window.read()
            print(f"Main event loop - Event: {event}")
            
            if event in (sg.WIN_CLOSED, "Exit"):
                break

            elif event == '-SETTINGS-':
                new_settings = open_settings_drawer(settings)
                if new_settings:
                    settings = new_settings
                    save_settings(settings)
                    sg.popup("Settings saved. Please restart the application for changes to take effect.")

            # ... (rest of your event handling code)

        window.close()

    except Exception as e:
        print(f"Error in main function: {str(e)}")
        traceback.print_exc()

    print("Exiting main function")

if __name__ == "__main__":
    main()
    print("Script completed")
    input("Press Enter to exit...")  # This will keep the console window open
    
    





























































































