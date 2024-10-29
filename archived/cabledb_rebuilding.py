import pandas as pd
import PySimpleGUI as sg
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import json
import traceback
import re
import numpy as np
import warnings
import time
import math
import base64
import logging
from logging.handlers import RotatingFileHandler
import sys
import tempfile
import win32com.client
import win32file
import win32con
import pywintypes
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows

# Define DEFAULT_REGEX_EXPRESSIONS globally
DEFAULT_REGEX_EXPRESSIONS = [
    ('(?i)pattern', 'Case-insensitive pattern'),
    ('pattern', 'Case-sensitive pattern'),
    (r'\d+', 'One or more digits'),
    (r'\w+', 'One or more word characters'),
    (r'\s+', 'One or more whitespace characters'),
    # Add more default patterns as needed
]

# Add this near the top of your file, after the imports
color_categories = [
    ("#FFFF00", "Yellow", "keyword1, keyword2"),
    ("#E6E6FA", "Lavender", "keyword3, keyword4"),
    ("#90EE90", "Light Green", "keyword5, keyword6"),
    ("#ADD8E6", "Light Blue", "keyword7, keyword8"),
    ("#FFB6C1", "Light Pink", "keyword9, keyword10"),
    ("#FFA500", "Orange", "keyword11, keyword12"),
    ("#00CED1", "Dark Turquoise", "keyword13, keyword14"),
    ("#FF69B4", "Hot Pink", "keyword15, keyword16")
]

# Set custom theme for loading animation
sg.LOOK_AND_FEEL_TABLE['LoadingTheme'] = {
    'BACKGROUND': '#F0F0F0',
    'TEXT': '#333333',
    'INPUT': '#FFFFFF',
    'TEXT_INPUT': '#000000',
    'SCROLL': '#99CC99',
    'BUTTON': ('#FFFFFF', '#333333'),
    'PROGRESS': ('#D0D0D0', '#FFFFFF'),
    'BORDER': 0, 'SLIDER_DEPTH': 0, 'PROGRESS_DEPTH': 0,
}

# Apply the custom theme
sg.theme('LoadingTheme')

print("Script started")

def add_new_records(columns, file_path):
    layout = [
        [sg.Text(col), sg.Input(key=f'-NEW-{col}-')] for col in columns
    ]
    layout.append([sg.Button('Add'), sg.Button('Cancel')])

    window = sg.Window('Add New Record', layout)

    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, 'Cancel'):
            window.close()
            return None
        elif event == 'Add':
            new_record = {col: values[f'-NEW-{col}-'] for col in columns}
            window.close()
            return new_record

    window.close()
    return None

def load_data(file_path):
    print(f"Loading data from {file_path}")
    try:
        df = pd.read_excel(file_path, sheet_name='CableList')
        length_matrix = pd.read_excel(file_path, sheet_name='LengthMatrix', index_col=0)
        print("Data loaded successfully")
        return df, length_matrix
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        raise
    except Exception as e:
        print(f"Error loading data: {str(e)}")
        raise

def apply_filter(df, values, columns_to_keep):
    filtered_df = df.copy()
    
    for col in columns_to_keep:
        if col == 'NUMBER':
            number_min = values.get(f"-FILTER-{col}-MIN-")
            number_max = values.get(f"-FILTER-{col}-MAX-")
            if number_min and number_max:
                try:
                    filtered_df = filtered_df[(filtered_df[col].astype(float) >= float(number_min)) & 
                                              (filtered_df[col].astype(float) <= float(number_max))]
                except ValueError:
                    logger.warning(f"Invalid {col} range")
                    sg.popup_error(f"Invalid {col} range")
        else:
            filter_value = values.get(f"-FILTER-{col}-")
            exact_match = values.get(f"-EXACT-{col}-", False)
            if filter_value:
                if exact_match:
                    filtered_df = filtered_df[filtered_df[col].astype(str).eq(filter_value)]
                else:
                    filtered_df = filtered_df[filtered_df[col].astype(str).str.contains(filter_value, case=False, na=False)]
    
    return filtered_df

def create_rack_mapping(length_matrix):
    rack_mapping = {}
    for rack in length_matrix.index:
        rack_mapping[rack] = rack
        if rack.startswith('T'):
            rack_mapping[rack[1:]] = rack  # Map "G01" to "TG01"
    return rack_mapping

def get_rack(name, mapping):
    return mapping.get(name[:4])  # Assume rack names are in the first 4 characters

def update_lengths_from_matrix(df, length_matrix, start_number, end_number, rack_mapping):
    print(f"Entering update_lengths_from_matrix. DataFrame shape: {df.shape}")
    mask = (df['NUMBER'] >= start_number) & (df['NUMBER'] <= end_number) & df['Length'].isna()
    df_to_update = df.loc[mask]
    print(f"Rows to update: {len(df_to_update)}")
    
    def get_length(row):
        origin_rack = get_rack(str(row['ORIGIN']), rack_mapping)
        dest_rack = get_rack(str(row['DEST']), rack_mapping)
        if origin_rack and dest_rack and origin_rack in length_matrix.index and dest_rack in length_matrix.columns:
            return length_matrix.loc[origin_rack, dest_rack]
        return np.nan
    
    df.loc[mask, 'Length'] = df_to_update.apply(get_length, axis=1)
    print(f"Updated DataFrame shape: {df.shape}")
    return df

def load_favorite_colors():
    try:
        with open('favorite_colors.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return ["#FFFF00", "#E6E6FA", "#90EE90", "#ADD8E6", "#FFB6C1", "#FFA500", "#00CED1", "#FF69B4"]

def save_favorite_colors(colors):
    with open('favorite_colors.json', 'w') as f:
        json.dump(colors, f)

def load_saved_regex():
    try:
        with open('saved_regex.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_regex(regex_dict):
    with open('saved_regex.json', 'w') as f:
        json.dump(regex_dict, f)

def update_table(window, df, settings):
    columns_to_hide = ['NUMC', 'Row'] + [f'F{i}' for i in range(11, 22)]
    visible_columns = [col for col in df.columns if col not in columns_to_hide]
    data = df[visible_columns].values.tolist()
    window["-TABLE-"].update(values=data)
    
    # Prepare display data
    display_data = []
    for _, row in df.iterrows():
        row_data = []
        for col in visible_columns:
            value = row[col]
            if pd.isna(value):
                row_data.append('')
            elif df[col].dtype == 'Int64':
                row_data.append(str(int(value)) if pd.notna(value) else '')
            else:
                row_data.append(str(value))
        display_data.append(row_data)
    
    print(f"Prepared display data. Rows: {len(display_data)}, Columns: {len(display_data[0]) if display_data else 0}")
    
    table = window["-TABLE-"]
    
    # Get current selections
    selected_rows = table.SelectedRows if hasattr(table, 'SelectedRows') else []
    print(f"Current selected rows: {selected_rows}")
    
    # Use default colors if not specified in settings
    primary_color = settings.get('primary_color', '#FFFFFF')
    secondary_color = settings.get('secondary_color', '#F0F0F0')
    text_color = settings.get('text_color', '#000000')
    
    # Create a list of row colors
    row_colors = [primary_color, secondary_color] * (len(display_data) // 2 + 1)
    row_color_list = [(i, text_color, color) for i, color in enumerate(row_colors[:len(display_data)])]
    
    print("Updating table...")
    # Update table values and colors
    table.update(values=display_data, row_colors=row_color_list)
    
    # Reapply row selection if any
    if selected_rows:
        print(f"Reapplying row selection: {selected_rows}")
        table.update(select_rows=selected_rows)
    
    window.refresh()
    print("Table updated and window refreshed")
    
def create_layout(df, length_matrix_headers, add_new_records, settings, file_path, columns_to_keep):
    # Filter layout
    filter_layout = []
    
    # Move "exact" label to the right
    filter_layout.append([sg.Push(), sg.Text("exact", size=(5, 1), font=('Helvetica', 8), 
                                  text_color=settings['function_text_color'], 
                                  background_color=settings['function_background_color'])])
    
    for col in columns_to_keep:
        row = [sg.Text(col, size=(10, 1), text_color=settings['function_text_color'], 
                       background_color=settings['function_background_color'])]
        if col == 'NUMBER':
            row.extend([
                sg.Input(key=f'-FILTER-{col}-MIN-', size=(7, 1), enable_events=True, 
                         background_color=settings['input_background_color']),
                sg.Text('-', background_color=settings['function_background_color']),
                sg.Input(key=f'-FILTER-{col}-MAX-', size=(7, 1), enable_events=True, 
                         background_color=settings['input_background_color']),
                sg.Push(),  # This will push the checkbox to the right
                sg.Checkbox('', key=f'-EXACT-{col}-', background_color=settings['function_background_color'])
            ])
        else:
            row.extend([
                sg.Input(key=f'-FILTER-{col}-', size=(20, 1), enable_events=True, 
                         background_color=settings['input_background_color']),
                sg.Push(),  # This will push the checkbox to the right
                sg.Checkbox('', key=f'-EXACT-{col}-', background_color=settings['function_background_color'])
            ])
        filter_layout.append(row)

    filter_layout.append([sg.Button("Apply Filter"), sg.Button("Clear Filter")])

    # Color layout
    color_layout = [
        [sg.Text("Color Categories", background_color=settings['function_background_color'], text_color=settings['function_text_color'])],
        [sg.Listbox(values=[f"{color} - {name}" for color, name, _ in color_categories], size=(20, 5), key='-COLOR-LIST-')],
        [sg.Input(key='-NEW-COLOR-', size=(10, 1)), sg.ColorChooserButton("Choose Color")],
        [sg.Input(key='-NEW-NAME-', size=(15, 1)), sg.Input(key='-NEW-KEYWORDS-', size=(15, 1))],
        [sg.Button("Add Category")]
    ]

    # Sort and Group layout
    sort_group_layout = [
        [sg.Text("Sort Column:", background_color=settings['function_background_color'], text_color=settings['function_text_color']),
         sg.Combo(df.columns.tolist(), key='-SORT-COLUMN-', size=(15, 1))],
        [sg.Radio("Ascending", "SORT", key='-SORT-ASCENDING-', default=True, background_color=settings['function_background_color'], text_color=settings['function_text_color']),
         sg.Radio("Descending", "SORT", key='-SORT-DESCENDING-', background_color=settings['function_background_color'], text_color=settings['function_text_color'])],
        [sg.Button("Sort")],
        [sg.Text("Group By:", background_color=settings['function_background_color'], text_color=settings['function_text_color'])],
        [sg.Radio("ORIGIN", "GROUP", key='-GROUP-ORIGIN-', default=True, background_color=settings['function_background_color'], text_color=settings['function_text_color']),
         sg.Radio("DEST", "GROUP", key='-GROUP-DEST-', background_color=settings['function_background_color'], text_color=settings['function_text_color'])],
        [sg.Input(key='-GROUP-VALUE-', size=(15, 1))],
        [sg.Button("Apply Grouping"), sg.Button("Reset Grouping")]
    ]

    # Action buttons
    action_buttons = [
        [sg.Button("LengthMatrix Lookup")],
        [sg.Button("Save Formatted Excel")],
        [sg.Button("Save Changes to Source")],
        [sg.Button("Add New Record")],
        [sg.Button("Import CSV")],
        [sg.Button("Reload Data")]
    ]

    # Combine all function layouts into a single column
    function_column = [
        sg.Column(filter_layout, background_color=settings['function_background_color']),
        sg.VSeparator(),
        sg.Column(color_layout, background_color=settings['function_background_color']),
        sg.VSeparator(),
        sg.Column(sort_group_layout, background_color=settings['function_background_color']),
        sg.VSeparator(),
        sg.Column(action_buttons, background_color=settings['function_background_color'])
    ]

    # Settings icon
    settings_icon = sg.Button('⚙️', key='-SETTINGS-', font=('Any', 20), button_color=(settings['text_color'], settings['background_color']), border_width=0)

    columns_to_hide = ['NUMC', 'Row'] + [f'F{i}' for i in range(11, 22)]
    visible_columns = [col for col in df.columns if col not in columns_to_hide]
    
    # Table layout
    table_layout = [
        [sg.Table(values=df[visible_columns].values.tolist(),
                  headings=visible_columns,
                  display_row_numbers=True,
                  auto_size_columns=False,
                  num_rows=25,
                  key="-TABLE-",
                  enable_events=True,
                  expand_x=True,
                  expand_y=True,
                  background_color=settings['background_color'],
                  text_color='black',
                  alternating_row_color='#F0F0F0',
                  header_background_color='#E0E0E0')]
    ]

    # Combine all layouts
    layout = [
        [settings_icon],
        [sg.Column([function_column], pad=(0,0), background_color=settings['function_background_color'])],
        [sg.HorizontalSeparator()],
        table_layout
    ]
    
    print("Layout created")
    return layout
    
def save_formatted_excel(df, color_categories, output_file):
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Convert NUMBER to string to remove decimal places
            df['NUMBER'] = df['NUMBER'].astype('Int64').astype(str)
            df.fillna('').to_excel(writer, sheet_name='Formatted_CableList', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Formatted_CableList']
            
            color_fills = {color_code: PatternFill(start_color=color_code.lstrip('#'), end_color=color_code.lstrip('#'), fill_type="solid") 
                           for color_code, _, _ in color_categories}
            
            for row in range(2, len(df) + 2):  # Start from 2 to skip header
                cell_color = worksheet.cell(row=row, column=df.columns.get_loc("Color") + 1).value
                fill = color_fills.get(cell_color)
                
                if fill:
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = fill
            
            worksheet.delete_cols(df.columns.get_loc("Color") + 1)  # Remove the Color column
        return True
    except PermissionError:
        sg.popup_error("Permission denied. The file may be open in another program.")
        return False
    except Exception as e:
        sg.popup_error(f"An error occurred while saving the file: {str(e)}")
        return False
        
   

   
        
def save_changes_to_excel(df, file_path):
    try:
        # Read the existing Excel file
        with pd.ExcelFile(file_path) as xls:
            other_sheets = {sheet_name: pd.read_excel(xls, sheet_name)
                            for sheet_name in xls.sheet_names if sheet_name != "CableList"}
        
        # Write the updated dataframe and other sheets back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="CableList", index=False)
            for sheet_name, sheet_df in other_sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return True
    except Exception as e:
        print(f"Error saving changes: {str(e)}")
        return False




def apply_sort(df, sort_column, ascending):
    return df.sort_values(by=sort_column, ascending=ascending)




def apply_grouping(df, group_by, group_value):
    if group_by not in ["ORIGIN", "DEST"]:
        return df
    
    column = group_by
    
    if not group_value:
        unique_values = df[column].unique()
        return pd.DataFrame({column: unique_values})
    
    mask = df[column].str.lower().str.contains(group_value.lower(), na=False)
    return df[mask]
    
    
    
    
def color_code_rows(df, color_categories):
    def get_color(row):
        origin_dest = f"{row['ORIGIN']} {row['DEST']}"
        for color_code, _, keywords in color_categories:
            if any(keyword.strip().lower() in origin_dest.lower() for keyword in keywords.split(',')):
                return color_code
        return "FFFFFF"  # White
    
    df["Color"] = df.apply(get_color, axis=1)
    return df





def generate_default_filename(df):
    current_date = datetime.now().strftime("%Y%m%d")
    dwg = df['DWG'].iloc[0] if 'DWG' in df.columns else 'UnknownDWG'
    project_id = df['Project ID'].iloc[0] if 'Project ID' in df.columns else 'UnknownProject'
    return f"{dwg}_{project_id}_{current_date}.xlsx"





def save_formatted_excel(df, file_path):
    try:
        output_path = sg.popup_get_file('Save Formatted Excel As', save_as=True, file_types=(("Excel Files", "*.xlsx"),))
        if output_path:
            df.to_excel(output_path, index=False)
            sg.popup(f"Formatted Excel saved to {output_path}")
    except Exception as e:
        sg.popup_error(f"Error saving formatted Excel: {str(e)}")

def save_changes_to_source(df, file_path):
    try:
        df.to_excel(file_path, index=False)
        sg.popup(f"Changes saved to source file: {file_path}")
    except Exception as e:
        sg.popup_error(f"Error saving changes to source: {str(e)}")

def import_csv(df):
    try:
        csv_path = sg.popup_get_file('Select CSV file to import', file_types=(("CSV Files", "*.csv"),))
        if csv_path:
            imported_df = pd.read_csv(csv_path)
            df = pd.concat([df, imported_df], ignore_index=True)
            sg.popup(f"CSV data imported from {csv_path}")
            return df
    except Exception as e:
        sg.popup_error(f"Error importing CSV: {str(e)}")
    return df




def manage_regex(column, current_regex, saved_regex):
    layout = [
        [sg.Text(f"Manage Regex for {column}")],
        [sg.Text("Current Regex:"), sg.Input(current_regex, key="-CURRENT_REGEX-", size=(30, 1))],
        [sg.Text("Saved Regex:"), sg.Listbox(list(saved_regex.get(column, {}).keys()), size=(30, 6), key="-SAVED_REGEX_LIST-")],
        [sg.Button("Save Current", bind_return_key=True), sg.Button("Apply Selected"), sg.Button("Delete Selected")],
        [sg.Button("Close")]
    ]
    
    window = sg.Window(f"Manage Regex for {column}", layout, finalize=True, return_keyboard_events=True)
    
    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Close", "\r"):  # \r is the Return key
            break
        elif event in ("Save Current", "\r"):  # Handle both button click and Return key
            regex = values["-CURRENT_REGEX-"]
            if regex:
                name = sg.popup_get_text(f"Enter a name for this {column} regex:")
                if name:
                    if column not in saved_regex:
                        saved_regex[column] = {}
                    saved_regex[column][name] = regex
                    save_regex(saved_regex)
                    window["-SAVED_REGEX_LIST-"].update(values=list(saved_regex[column].keys()))
                    print(f"Saved regex '{name}' for column '{column}': {regex}")
        elif event == "Apply Selected":
            selected = values["-SAVED_REGEX_LIST-"]
            if selected:
                selected_regex = selected[0]
                regex = saved_regex[column][selected_regex]
                window["-CURRENT_REGEX-"].update(regex)
                print(f"Applied regex '{selected_regex}' to column '{column}': {regex}")
        elif event == "Delete Selected":
            selected = values["-SAVED_REGEX_LIST-"]
            if selected:
                selected_regex = selected[0]
                del saved_regex[column][selected_regex]
                save_regex(saved_regex)
                window["-SAVED_REGEX_LIST-"].update(values=list(saved_regex[column].keys()))
                print(f"Deleted regex '{selected_regex}' from column '{column}'")
    
    window.close()
    return values["-CURRENT_REGEX-"]





def handle_length_matrix_lookup(df, length_matrix, window, rack_mapping):
    layout = [
        [sg.Text("Select range of cable numbers to update:")],
        [sg.Text("Start Number:", size=(12, 1), justification='right'), sg.Input(key="-START_NUMBER-", size=(6, 1))],
        [sg.Text("End Number:", size=(12, 1), justification='right'), sg.Input(key="-END_NUMBER-", size=(6, 1))],
        [sg.Button("Preview", bind_return_key=True), sg.Button("Cancel")]
    ]
    
    range_window = sg.Window("LengthMatrix Lookup", layout, finalize=True, return_keyboard_events=True)
    
    while True:
        event, values = range_window.read()
        print(f"LengthMatrix Lookup event: {event}")
        
        if event in (sg.WIN_CLOSED, "Cancel", "\r"):
            print("Exiting LengthMatrix Lookup without changes")
            break
        
        if event in ("Preview", "\r"):
            try:
                start_number = int(values["-START_NUMBER-"])
                end_number = int(values["-END_NUMBER-"])
                print(f"Start number: {start_number}, End number: {end_number}")
                
                if start_number > end_number:
                    raise ValueError("Invalid number range")
                
                preview_df = update_lengths_from_matrix(df.copy(), length_matrix, start_number, end_number, rack_mapping)
                print(f"Preview dataframe created. Shape: {preview_df.shape}")
                
                preview_mask = (preview_df['NUMBER'] >= start_number) & (preview_df['NUMBER'] <= end_number)
                preview_data = preview_df[preview_mask]
                print(f"Preview data created. Shape: {preview_data.shape}")

                preview_data_str = preview_data.astype(str).replace('nan', '')
                
                preview_layout = [
                    [sg.Table(values=preview_data_str.values.tolist(),
                              headings=preview_data.columns.tolist(),
                              display_row_numbers=False,
                              auto_size_columns=False,
                              num_rows=min(10, len(preview_data)),
                              key="-PREVIEW_TABLE-")],
                    [sg.Button("Confirm"), sg.Button("Cancel")]
                ]
                preview_window = sg.Window("Preview Changes", preview_layout)
                
                preview_event, _ = preview_window.read()
                print(f"Preview window event: {preview_event}")
                if preview_event == "Confirm":
                    df = update_lengths_from_matrix(df, length_matrix, start_number, end_number, rack_mapping)
                    print("LengthMatrix lookup completed successfully")
                    sg.popup("LengthMatrix lookup completed successfully!")
                    preview_window.close()
                    range_window.close()
                    return df
                
                preview_window.close()
            
            except ValueError as e:
                print(f"ValueError in LengthMatrix Lookup: {str(e)}")
                sg.popup_error(f"Error: {str(e)}")
            except Exception as e:
                print(f"Unexpected error in LengthMatrix Lookup: {str(e)}")
                sg.popup_error(f"An unexpected error occurred: {str(e)}")
    
    range_window.close()
    print("Exiting handle_length_matrix_lookup function")
    return df
    
    
    
    
    
  

















  

def create_circle_animation(canvas, center, radius, color):
    start_angle = 0
    extent = 270
    canvas.TKCanvas.create_arc(center[0]-radius, center[1]-radius,
                               center[0]+radius, center[1]+radius,
                               start=start_angle, extent=extent,
                               outline=color, width=4, style='arc')

def rotate_circle_animation(canvas, center, radius, color, angle):
    canvas.TKCanvas.delete('all')
    start_angle = angle % 360
    extent = 270
    canvas.TKCanvas.create_arc(center[0]-radius, center[1]-radius,
                               center[0]+radius, center[1]+radius,
                               start=start_angle, extent=extent,
                               outline=color, width=4, style='arc')

def create_loading_animation(canvas, center, radius, color, thickness):
    canvas.TKCanvas.create_oval(
        center[0] - radius, center[1] - radius,
        center[0] + radius, center[1] + radius,
        outline=color, width=thickness
    )

def update_loading_animation(canvas, center, radius, color, thickness, angle):
    canvas.TKCanvas.delete('all')
    start_angle = math.radians(angle)
    end_angle = start_angle + math.pi * 1.5  # Extend the arc to 270 degrees
    
    canvas.TKCanvas.create_arc(
        center[0] - radius, center[1] - radius,
        center[0] + radius, center[1] + radius,
        start=angle, extent=270, outline=color, width=thickness, style='arc'
    )

def show_loading_animation():
    layout = [[sg.Canvas(size=(100, 100), key='-CANVAS-', background_color='#F8F8F8')]]
    window = sg.Window('Loading', layout, finalize=True, keep_on_top=True, 
                       no_titlebar=True, alpha_channel=0.2,  # Increased transparency
                       background_color='#F8F8F8')
    
    canvas = window['-CANVAS-']
    center = (50, 50)
    radius = 40
    color = '#0077BE'  # ABC7 blue
    thickness = 6  # Thicker line for a more modern look
    
    create_loading_animation(canvas, center, radius, color, thickness)
    
    angle = 0
    start_time = time.time()
    while time.time() - start_time < 8:  # Run for  seconds
        event, values = window.read(timeout=20)
        if event == sg.WINDOW_CLOSED:
            break
        angle += 5
        update_loading_animation(canvas, center, radius, color, thickness, angle)
    
    window.close()

# ABC7 logo (replace this with the actual base64 encoded ABC7 logo)

def get_img_data(b64str, maxsize=(100, 50), first=False):
    b64_bytes = base64.b64decode(b64str)
    import io
    from PIL import Image
    im = Image.open(io.BytesIO(b64_bytes))
    im.thumbnail(maxsize)
    if first:                     
        im.thumbnail((maxsize[0] // 2, maxsize[1] // 2))
    with io.BytesIO() as output:
        im.save(output, format="PNG")
        data = output.getvalue()
    return data

def open_settings_drawer(current_settings):
    light_theme = {
        'background_color': '#F5F5F5',  # Light gray inspired by ABC7
        'text_color': '#1A1A1A',        # Dark gray
        'button_color': ('#FFFFFF', '#0077BE'),  # White text, ABC7 blue
        'input_background_color': '#FFFFFF',  # White
        'function_background_color': '#E0E0E0',  # Lighter gray for function sections
        'function_text_color': '#000000'  # Black text for function sections
    }
    
    dark_theme = {
        'background_color': '#1A1A1A',  # Dark gray
        'text_color': '#F5F5F5',        # Light gray
        'button_color': ('#FFFFFF', '#0077BE'),  # White text, ABC7 blue
        'input_background_color': '#2C2C2C',  # Darker gray
        'function_background_color': '#2C2C2C',  # Same darker gray for function sections
        'function_text_color': '#FFFFFF'  # White text for function sections
    }

    layout = [
        [sg.Text("Settings", font=('Helvetica', 16), pad=(0, 10))],
        [sg.Frame('Theme', [
            [sg.Radio("Light", "THEME", key="-LIGHT-", default=True), 
             sg.Radio("Dark", "THEME", key="-DARK-")]
        ], relief=sg.RELIEF_FLAT, pad=(0, 10))],
        [sg.Frame('Color Settings', [
            [sg.Text("Background:", size=(12, 1)), 
             sg.Input(current_settings['background_color'], key='-BACKGROUND_COLOR-', size=(10, 1)),
             sg.ColorChooserButton("", target='-BACKGROUND_COLOR-', button_color=('white', current_settings['background_color']), size=(2, 1))],
            [sg.Text("Text:", size=(12, 1)), 
             sg.Input(current_settings['text_color'], key='-TEXT_COLOR-', size=(10, 1)),
             sg.ColorChooserButton("", target='-TEXT_COLOR-', button_color=('white', current_settings['text_color']), size=(2, 1))],
            [sg.Text("Button:", size=(12, 1)), 
             sg.Input(current_settings['button_color'][1], key='-BUTTON_COLOR-', size=(10, 1)),
             sg.ColorChooserButton("", target='-BUTTON_COLOR-', button_color=('white', current_settings['button_color'][1]), size=(2, 1))],
            [sg.Text("Input Background:", size=(12, 1)), 
             sg.Input(current_settings['input_background_color'], key='-INPUT_BACKGROUND_COLOR-', size=(10, 1)),
             sg.ColorChooserButton("", target='-INPUT_BACKGROUND_COLOR-', button_color=('white', current_settings['input_background_color']), size=(2, 1))],
            [sg.Text("Function Background:", size=(12, 1)), 
             sg.Input(current_settings['function_background_color'], key='-FUNCTION_BACKGROUND_COLOR-', size=(10, 1)),
             sg.ColorChooserButton("", target='-FUNCTION_BACKGROUND_COLOR-', button_color=('white', current_settings['function_background_color']), size=(2, 1))],
            [sg.Text("Function Text:", size=(12, 1)), 
             sg.Input(current_settings['function_text_color'], key='-FUNCTION_TEXT_COLOR-', size=(10, 1)),
             sg.ColorChooserButton("", target='-FUNCTION_TEXT_COLOR-', button_color=('white', current_settings['function_text_color']), size=(2, 1))],
        ], relief=sg.RELIEF_FLAT, pad=(0, 10))],
        [sg.Button("Save", size=(8, 1)), sg.Button("Reset to Defaults", size=(15, 1)), sg.Button("Cancel", size=(8, 1))]
    ]

    window = sg.Window("Settings", layout, finalize=True, keep_on_top=True, use_default_focus=False)

    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, "Cancel"):
            break
        if event == "Save":
            new_settings = {
                'background_color': values['-BACKGROUND_COLOR-'],
                'text_color': values['-TEXT_COLOR-'],
                'button_color': (current_settings['button_color'][0], values['-BUTTON_COLOR-']),
                'input_background_color': values['-INPUT_BACKGROUND_COLOR-'],
                'function_background_color': values['-FUNCTION_BACKGROUND_COLOR-'],
                'function_text_color': values['-FUNCTION_TEXT_COLOR-'],
                'theme': 'light' if values['-LIGHT-'] else 'dark'
            }
            window.close()
            return new_settings
        if event == "Reset to Defaults":
            theme = light_theme if values['-LIGHT-'] else dark_theme
            for key, value in theme.items():
                if isinstance(value, tuple):
                    window[f'-{key.upper()}-'].update(value[1])
                else:
                    window[f'-{key.upper()}-'].update(value)
        if event in ('-LIGHT-', '-DARK-'):
            theme = light_theme if values['-LIGHT-'] else dark_theme
            for key, value in theme.items():
                if isinstance(value, tuple):
                    window[f'-{key.upper()}-'].update(value[1])
                else:
                    window[f'-{key.upper()}-'].update(value)

    window.close()
    return None

def load_last_file_path():
    try:
        with open('last_file_path.json', 'r') as f:
            data = json.load(f)
            return data.get('last_path', '')
    except FileNotFoundError:
        return ''
    except json.JSONDecodeError:
        print("Error decoding last_file_path.json. Using empty path.")
        return ''

def save_last_file_path(file_path):
    with open('last_file_path.json', 'w') as f:
        json.dump({'last_path': file_path}, f)

def load_settings():
    try:
        with open('settings.json', 'r') as f:
            settings = json.load(f)
    except FileNotFoundError:
        settings = {'theme': 'dark', 'font_size': 12, 'window_size': (800, 600), 'window_location': (None, None), 'projectid_required': True}
    return settings

def save_settings(settings):
    with open('settings.json', 'w') as f:
        json.dump(settings, f)

def is_file_accessible(file_path, mode='r'):
    try:
        file = open(file_path, mode)
        file.close()
        return True
    except IOError:
        return False

def load_excel_file(file_path=None):
    logger = logging.getLogger('CableDB')
    logger.info("Entering load_excel_file function")
    
    if file_path is None:
        file_path = sg.popup_get_file("Select Excel file to load", 
                                      file_types=(("Excel Files", "*.xlsx;*.xlsm"),))
        if not file_path:
            return None, None, None, None
    
    logger.info(f"Attempting to load data from {file_path}")
    
    try:
        # Try to create a temporary copy
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{os.path.basename(file_path)}")
        shutil.copy2(file_path, temp_file)
        
        # Read from the temporary file
        df = pd.read_excel(temp_file, sheet_name="CableList", dtype=object)
        length_matrix = pd.read_excel(temp_file, sheet_name="LengthMatrix", dtype=object)
        columns_to_keep = df.columns.tolist()
        
        return df, length_matrix, temp_file, columns_to_keep
    except PermissionError:
        logger.error(f"Permission denied when trying to copy {file_path}")
        sg.popup_error(f"Permission denied when trying to access {file_path}. The file might be open in another program.")
        return None, None, None, None
    except Exception as e:
        logger.error(f"Error loading file: {str(e)}")
        sg.popup_error(f"An error occurred while loading the file:\n{str(e)}")
        return None, None, None, None

def setup_logging():
    # Create a logger
    logger = logging.getLogger('CableDB')
    logger.setLevel(logging.DEBUG)

    # Create handlers
    normal_handler = RotatingFileHandler('normal.log', maxBytes=1000000, backupCount=5)
    normal_handler.setLevel(logging.INFO)

    crash_handler = RotatingFileHandler('crash.log', maxBytes=1000000, backupCount=5)
    crash_handler.setLevel(logging.ERROR)

    # Create formatters and add it to handlers
    normal_format = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    crash_format = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s\n%(exc_info)s')
    
    normal_handler.setFormatter(normal_format)
    crash_handler.setFormatter(crash_format)

    # Add handlers to the logger
    logger.addHandler(normal_handler)
    logger.addHandler(crash_handler)

    return logger

def add_new_records_dialog(columns_to_keep, settings):
    logger = logging.getLogger('CableDB')
    logger.info("Opening Add New Records dialog")

    field_lengths = {
        'NUMBER': 10,
        'DWG': 15,
        'ORIGIN': 20,
        'DEST': 20,
        'Alternate Dwg': 15,
        'Wire Type': 15,
        'Length': 10,
        'Note': 30,
        'ProjectID': 15
    }

    layout = [
        [sg.Text("Add New Records", font=("Helvetica", 16), justification='center', expand_x=True)],
        [sg.HorizontalSeparator()],
    ]

    for col in columns_to_keep:
        layout.append([
            sg.Text(col, size=(15, 1), justification='right'),
            sg.Input(key=f'-NEW-{col}-', size=(field_lengths.get(col, 20), 1), enable_events=True)
        ])

    layout.extend([
        [sg.HorizontalSeparator()],
        [sg.Button('Add Record', key='-ADD-', disabled=True), 
         sg.Button('Clear Fields', key='-CLEAR-'), 
         sg.Button('Done', key='-DONE-')]
    ])

    window = sg.Window('Add New Records', layout, finalize=True, modal=True, return_keyboard_events=True)
    
    new_records = []

    def update_add_button():
        required_fields = ['NUMBER', 'DWG', 'ORIGIN', 'DEST']
        if settings.get('projectid_required', True):
            required_fields.append('ProjectID')
        values = window.read(timeout=0)[1]
        all_required_filled = all(values[f'-NEW-{field}-'].strip() for field in required_fields if f'-NEW-{field}-' in values)
        window['-ADD-'].update(disabled=not all_required_filled)
        if all_required_filled:
            window['-ADD-'].set_focus()

    while True:
        event, values = window.read()
        
        if event in (sg.WIN_CLOSED, '-DONE-'):
            break
        
        elif event.startswith('-NEW-'):
            update_add_button()
        
        elif event == '-ADD-' or (event == '\r' and not window['-ADD-'].Disabled):
            new_record = {col: values[f'-NEW-{col}-'] for col in columns_to_keep}
            new_records.append(new_record)
            logger.info(f"New record added: {new_record}")
            sg.popup_quick_message("Record added successfully!", background_color='green', text_color='white')
            
            # Clear fields after adding
            for col in columns_to_keep:
                window[f'-NEW-{col}-'].update('')
            update_add_button()
        
        elif event == '-CLEAR-':
            for col in columns_to_keep:
                window[f'-NEW-{col}-'].update('')
            update_add_button()

    window.close()
    return new_records

def is_file_open(file_path):
    if os.path.exists(file_path):
        try:
            os.rename(file_path, file_path)
            return False
        except OSError:
            return True
    return False

def save_to_excel(df, original_file_path, sheet_name='CableList'):
    logger = logging.getLogger('CableDB')
    
    # Generate a new file name with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    file_dir = os.path.dirname(original_file_path)
    file_name = os.path.basename(original_file_path)
    name, ext = os.path.splitext(file_name)
    new_file_path = os.path.join(file_dir, f"{name}_{timestamp}{ext}")
    
    try:
        # First, create a copy of the original file
        shutil.copy2(original_file_path, new_file_path)
        
        # Now, open the new file and replace the CableList sheet
        with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.book.remove(writer.book[sheet_name])
            writer.book.active = writer.book[sheet_name]
        
        logger.info(f"Data saved successfully to {new_file_path}")
        return new_file_path
    except Exception as e:
        logger.error(f"Error saving file: {str(e)}")
        sg.popup_error(f"An error occurred while saving the file:\n{str(e)}")
        return None

def check_file_accessibility(file_path):
    try:
        # Try to open the file with write access
        handle = win32file.CreateFile(
            file_path,
            win32con.GENERIC_WRITE,
            0,  # Exclusive access
            None,
            win32con.OPEN_EXISTING,
            win32con.FILE_ATTRIBUTE_NORMAL,
            None
        )
        win32file.CloseHandle(handle)
        return True, None
    except pywintypes.error as e:
        if e.winerror == 32:
            return False, "File is open in another program"
        elif e.winerror == 5:
            return False, "You don't have permission to modify this file"
        else:
            return False, f"Error accessing file: {str(e)}"

def create_backup(file_path):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{file_path[:-5]}_backup_{timestamp}.xlsm"
    try:
        shutil.copy2(file_path, backup_path)
        return backup_path
    except Exception as e:
        logger.error(f"Failed to create backup: {str(e)}")
        return None

def create_working_copy(file_path):
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, f"temp_{os.path.basename(file_path)}")
    shutil.copy2(file_path, temp_file)
    return temp_file

def main():
    logger = setup_logging()
    logger.info("Script started")
    try:
        show_loading_animation()  # Show loading animation at start for 6 seconds

        # Load settings
        settings = load_settings()
        theme = settings['theme']
        
        if theme == 'light':
            sg.theme('LightGrey1')
        else:
            sg.theme('DarkGrey9')
        
        background_color = settings['background_color']
        text_color = settings['text_color']
        button_color = settings['button_color']
        input_background_color = settings['input_background_color']

        # Start with empty DataFrames
        df = pd.DataFrame()
        length_matrix = pd.DataFrame()
        file_path = None
        temp_file = None
        columns_to_keep = []

        # Create layout and window
        layout = create_layout(df, length_matrix.columns.tolist(), add_new_records, settings, file_path, columns_to_keep)
        window = sg.Window("Cable Database Interface", layout, resizable=True, finalize=True, 
                           size=settings['window_size'], location=settings['window_location'])

        # Event Loop
        while True:
            event, values = window.read()
            logger.debug(f"Event: {event}")

            if event in (sg.WIN_CLOSED, "Exit"):
                # Save window size and location before closing
                settings['window_size'] = window.size
                settings['window_location'] = window.current_location()
                save_settings(settings)
                break

            elif event == "Apply Filter":
                df_filtered = apply_filter(df, values, columns_to_keep)
                update_table(window, df_filtered, settings)

            elif event == "Clear Filter":
                df_filtered = clear_filter(df)
                update_table(window, df_filtered, settings)

            elif event == "LengthMatrix Lookup":
                # Implement LengthMatrix Lookup functionality
                sg.popup("LengthMatrix Lookup functionality not implemented yet.")

            elif event == "Save Formatted Excel":
                # Implement Save Formatted Excel functionality
                save_formatted_excel(df, file_path)

            elif event == "Save Changes to Source":
                # Implement Save Changes to Source functionality
                save_changes_to_source(df, file_path)

            elif event == "Add New Record":
                if df.empty:
                    sg.popup_error("Please load data first before adding new records.")
                    continue
                
                new_records = add_new_records_dialog(columns_to_keep, settings)
                if new_records:
                    # Convert the new records to a DataFrame
                    new_df = pd.DataFrame(new_records)
                    for col in df.columns:
                        new_df[col] = new_df[col].astype(df[col].dtype)
                    
                    # Append the new records to the existing DataFrame
                    df = pd.concat([df, new_df], ignore_index=True)
                    
                    # Save to a new Excel file
                    new_file_path = save_to_excel(df, file_path)
                    if new_file_path:
                        file_path = new_file_path  # Update the current file path
                        update_table(window, df, settings)
                        logger.info(f"{len(new_records)} new records added and saved to new Excel file")
                        sg.popup(f"{len(new_records)} new records added successfully!\n"
                                 f"Saved to new file: {new_file_path}")
                    else:
                        logger.warning("Failed to save to a new Excel file")
                        sg.popup_ok("Failed to save to a new Excel file. "
                                    "Please check the application log for details.")

            elif event == "Import CSV":
                # Implement Import CSV functionality
                import_csv(df)

            elif event == "Reload Data":
                # Ask user to select the file to load
                new_file_path = sg.popup_get_file("Select Excel file to load", 
                                                 default_path=file_path, 
                                                 file_types=(("Excel Files", "*.xlsx;*.xlsm"),))
                if new_file_path:
                    df, length_matrix, temp_file, columns_to_keep = load_excel_file(new_file_path)
                    if df is not None and length_matrix is not None:
                        file_path = new_file_path
                        update_table(window, df, settings)
                        logger.info("Data reloaded successfully")
                        sg.popup("Data reloaded successfully!")

            elif event == "Sort":
                sort_column = values['-SORT-COLUMN-']
                ascending = values['-SORT-ASCENDING-']
                if sort_column:
                    df = df.sort_values(by=sort_column, ascending=ascending)
                    update_table(window, df, settings)

            elif event == "Apply Grouping":
                group_by = 'ORIGIN' if values['-GROUP-ORIGIN-'] else 'DEST'
                group_value = values['-GROUP-VALUE-']
                if group_value:
                    df_grouped = df[df[group_by] == group_value]
                    update_table(window, df_grouped, settings)

            elif event == "Reset Grouping":
                update_table(window, df, settings)

            elif event == "Settings":
                settings_window = create_settings_window(settings)
                settings_event, new_settings = settings_window.read(close=True)
                if settings_event == 'Save':
                    settings['theme'] = new_settings['-THEME-']
                    settings['font_size'] = new_settings['-FONT_SIZE-']
                    settings['projectid_required'] = new_settings['-PROJECTID_REQUIRED-']
                    save_settings(settings)
                    sg.popup('Settings saved. Please restart the application for changes to take effect.')

    except Exception as e:
        logger.exception("An unexpected error occurred in the main function")
        sg.popup_error(f"An unexpected error occurred: {str(e)}\nPlease check the crash log for details.")
    finally:
        logger.info("Script completed")

if __name__ == "__main__":
    logger = setup_logging()
    try:
        main()
    except Exception as e:
        logger.critical("An unhandled exception occurred", exc_info=True)
        sg.popup_error(f"A critical error occurred: {str(e)}\nPlease check the crash log for details.")

